from openpyxl import load_workbook

filepath = "files/wildberries.xlsx"

workbook = load_workbook(filepath)

sheet = workbook[workbook.sheetnames[0]]


def get_sku():
    for row in sheet.iter_rows(values_only=True):
        yield row[0]
